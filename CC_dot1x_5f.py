"""
Python script by Hardik Raval
If any error/issues, please reach out to haraval@cisco.com with Subject "<Customer> - CC-dot1x-Status"
This script lists switch access interfaces and their Dot1x configuration status.
"""

import requests
import time
import urllib3
import pandas as pd
import os
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Load environment variables from .env file
load_dotenv()

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- Constants ---
DNAC_IP = os.getenv("DNAC_IP")
USERNAME = os.getenv("USERNAME")
PASSWORD = os.getenv("PASSWORD")
BASE_URL = f"https://{DNAC_IP}/dna"
VERIFY_SSL = False

def print_commentary(msg):
    print(f"==> {msg}")

# --- Authentication: Get Token ---
def get_auth_token():
    url = f"{BASE_URL}/system/api/v1/auth/token"
    print_commentary("Requesting authentication token from Catalyst Center...")
    response = requests.post(url, auth=(USERNAME, PASSWORD), verify=VERIFY_SSL)
    response.raise_for_status()
    print_commentary("Authentication successful.")
    return response.json()["Token"]

def get_all_devices(token):
    url = f"{BASE_URL}/intent/api/v1/network-device"
    headers = {"X-Auth-Token": token}
    print_commentary("Fetching all network devices from inventory...")
    response = requests.get(url, headers=headers, verify=VERIFY_SSL)
    response.raise_for_status()
    devices = response.json().get("response", [])
    print_commentary(f"Total devices found: {len(devices)}")
    return devices

def is_dot1x_globally_enabled(token, device_id, max_retries=5):
    url = f"{BASE_URL}/intent/api/v1/wired/networkDevices/{device_id}/configFeatures/deployed/layer2/dot1xGlobalConfig"
    headers = {"X-Auth-Token": token}
    retries = 0
    while retries < max_retries:
        response = requests.get(url, headers=headers, verify=VERIFY_SSL)
        if response.status_code == 429:
            retry_after = int(response.headers.get("Retry-After", "60"))
            print_commentary(f"Rate limit hit (HTTP 429). Waiting for {retry_after} seconds before retrying...")
            time.sleep(retry_after)
            retries += 1
            continue
        try:
            response.raise_for_status()
        except Exception as e:
            print_commentary(f"Error HTTP {response.status_code} during dot1xGlobal check for device {device_id}: {response.text}")
            return False
        api_data = response.json()
        global_config = api_data.get('response', {}).get('dot1xGlobalConfig', {}).get('items', [])
        return global_config and global_config[0].get('isDot1xEnabled', False)
    print_commentary(f"Max retries exceeded for device {device_id}. Skipping.")
    return False


def get_all_interfaces(token, device_id):
    url = f"{BASE_URL}/intent/api/v1/interface/network-device/{device_id}"
    headers = {"X-Auth-Token": token}
    response = requests.get(url, headers=headers, verify=VERIFY_SSL)
    response.raise_for_status()
    return response.json().get('response', [])

def get_dot1x_interfaces(token, device_id):
    url = f"{BASE_URL}/intent/api/v1/wired/networkDevices/{device_id}/configFeatures/deployed/layer2/dot1xInterfaceConfig"
    headers = {"X-Auth-Token": token}
    response = requests.get(url, headers=headers, verify=VERIFY_SSL)
    response.raise_for_status()
    api_data = response.json()
    interface_configs = api_data.get('response', {}).get('dot1xInterfaceConfig', {}).get('items', [])
    return [
        iface.get('interfaceName', "Unknown")
        for iface in interface_configs
        if iface.get('configType') == 'DOT1X_INTERFACE'
    ]

def format_sheet(ws):
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align
            cell.border = thin_border
            if row == 1:
                cell.font = header_font
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

def main():
    token = get_auth_token()
    devices = get_all_devices(token)
    output_rows = []

    switches = [
        d for d in devices
        if isinstance(d.get("type"), str) and "Switch" in d["type"]
    ]
    print_commentary(f"Total switches in inventory: {len(switches)}")

    for idx, switch in enumerate(switches, 1):
        device_id = switch["id"]
        hostname = switch.get("hostname", "N/A")
        print_commentary(f"[{idx}/{len(switches)}] Checking switch '{hostname}' (ID: {device_id}) for Dot1x global status...")
        dot1x_global = is_dot1x_globally_enabled(token, device_id)
        if dot1x_global:
            print_commentary(f"Dot1x is globally enabled on {hostname}. Gathering interface data...")
            all_interfaces = get_all_interfaces(token, device_id)
            dot1x_interfaces = set(get_dot1x_interfaces(token, device_id))

            for iface in all_interfaces:
                iface_name = iface.get("portName", iface.get("interfaceName", "Unknown"))
                if iface_name == "Unknown":
                    continue
                port_mode = iface.get("portMode") or iface.get("interfaceMode")
                if str(port_mode).lower() != "access":
                    continue
                has_dot1x = "Yes" if iface_name in dot1x_interfaces else "No"
                output_rows.append({
                    "Device Name": hostname,
                    "Device ID": device_id,
                    "Interface Name": iface_name,
                    "Dot1x Configured": has_dot1x
                })
        else:
            print_commentary(f"Dot1x is NOT globally enabled on {hostname}. Skipping interface check.")

    if not output_rows:
        print_commentary("No access interfaces found with/without Dot1x configuration. Exiting.")
        return

    # --- Export to Excel ---
    df = pd.DataFrame(output_rows)
    output_file = "dot1x_access_interface_report.xlsx"
    print_commentary("Exporting results to Excel...")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Dot1x Access Interfaces", index=False)
        #writer.save()

    # Open with openpyxl for formatting and chart
    wb = load_workbook(output_file)
    ws = wb["Dot1x Access Interfaces"]
    format_sheet(ws)

    # --- Add Summary Sheet and Bar Chart ---
    print_commentary("Generating summary and bar graph...")
    summary = df["Dot1x Configured"].value_counts().rename_axis('Dot1x Configured').reset_index(name='Interface Count')
    ws2 = wb.create_sheet("Dot1x Summary")
    ws2.append(["Dot1x Configured", "Interface Count"])
    for row in summary.itertuples(index=False):
        ws2.append(list(row))
    format_sheet(ws2)

    # Create a bar chart
    chart = BarChart()
    chart.title = "Access Interfaces with/without Dot1x"
    chart.x_axis.title = "Dot1x Configured"
    chart.y_axis.title = "Interface Count"
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 7
    ws2.add_chart(chart, "E2")

    wb.save(output_file)
    print_commentary(f"Report exported to {output_file}")

if __name__ == "__main__":
    main()